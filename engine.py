import os
import openpyxl
from openpyxl import load_workbook


#load the workbook, and chose the active sheet
workbook=load_workbook(filename=r"C:\Users\nkurien\Desktop\Copier Project\Final\Copiers.xlsx").active

count=workbook.max_row 

#the rows with IP's starts at 2
start =2


while( start <= count):
    beginning= start
    for i in range(10):
        ipAddress=workbook["E"+str(start)].value
        if ipAddress==None:
            start=start+1
            pass
        else:
            os.system("start http://"+ipAddress)
            print("Serial # is "+str(workbook["B"+str(start)].value))
            print("Serial # is "+str(workbook["C"+str(start)].value))
            input(str(start)+"-> to the next")
            try:
                start=start+1
            except Exception:
                print("got to the end")
    
    input("Finished {0} to {1}, Ready to go to the next 10".format(beginning, start))



